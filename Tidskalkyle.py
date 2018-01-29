# -*- coding: utf-8 -*-
"""
Created on Fri Jan 19 10:01:03 2018

@author: Heiv085
"""

#from openpyxl import load_workbook
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os


class TimeCalc():
    def __init__(self, master):
        
        self.master = master
        self.hour_hold = 0
        self.minute_hold = 0
        self.second_hold = 0
        
        self.tool_cnt = 0
        self.cs_cnt = 0
        
        self.kill_task()
        
        self.wb = openpyxl.load_workbook(filename=self.master)
        
        self.readworkbook()
        self.pnumber()
                
    def readworkbook(self):
        
        """ Reading through the generated report file from GibbsCam
            and passing on the time, tool and coordinate system to the 
            timemagic method """
        
        ws = self.wb.active
        
        self.length_counter = 11
        run_time = 12
        cs_plane = 8
        
        while ws.cell(row=self.length_counter, column=4).value != None:
            
            tool = ws.cell(row=self.length_counter, column=4).value
            time = ws.cell(row=run_time, column=10).value
            cs = ws.cell(row=cs_plane, column=8).value
            self.timemagic(time, tool, cs)
            
            self.length_counter += 8
            run_time += 8
            cs_plane += 8
        
    def timemagic(self, time, tool, cs):
        
        """ Taking each toolpath time and adds toolchanging times and idle time
            to them to get a better estiamte of machining time """
        
        time = time
        tool = tool
        tool = int(tool)
        
        cs = cs
        cs = int(cs)
        
        hour = time[:1]
        hour = int(hour)
        
        minute = time[2:4]
        minute = int(minute)
        
        second = time[5:]
        second = float(second)
        
        if self.length_counter == 10:
            self.second_hold += 8
        
        
        # baseline converting the input and adding them to separate holding
        # variables.
        if hour == 0:
            pass
        else:
            self.hour_hold += hour
            
        
        if self.minute_hold > 59:
            self.hour_hold += 1
            self.minute_hold -= 59
            
            
        if minute == 0:
            pass
        elif minute > 59:
            self.hour_hold += 1
        else:
            self.minute_hold += minute
            
        
        while self.second_hold > 60:
            self.minute_hold += 1
            self.second_hold -= 60
            
            
        if second == 0:
            pass
        elif second > 60:
            self.minute_hold += 1
        else:
            self.second_hold += second
        
        
        second_tmp = 0
        
        # Decide if the same tool is used over and if add flatt 5 sec to the 
        # counter, if not check time of toolpath and decide from there
        
        if tool == self.tool_cnt:
            self.second_hold += 5
        else:
            if minute > 0:
                second_tmp += minute * 60
                second_tmp += second
            else:
                second_tmp += second
                
            if second_tmp > 5:
                self.second_hold += 16.0
                
            elif second_tmp <= 5:
                self.second_hold += 5 % second_tmp + 16
                
        if cs > 1 and cs != self.cs_cnt:
            self.second_hold += 7
               
        self.tool_cnt = tool
        self.cs_cnt = cs
        
        while self.second_hold > 60:
            self.minute_hold += 1
            self.second_hold -= 60
        
    def results(self):
        
        """ Converts the results to strings and returnes them """
        
        self.second_hold = round(self.second_hold, 2)
        
        self.hour_hold = str(self.hour_hold)
        self.minute_hold = str(self.minute_hold)
        self.second_hold = str(self.second_hold)
        
        result = self.hour_hold + ':' + self.minute_hold + ':' + self.second_hold
        
        #os.remove(self.master)
        
        return result
        
    def pnumber(self):
        
        """ Retrives and returnes the part code """
        
        ws = self.wb.active
        p_number = ws.cell(row=4, column=10).value
        #p_number = p_number[:6]
        #p_number = int(p_number)
        
        return p_number
        
    def kill_task(self):
        
        """ Killing every open instance of excel """
        
        os.system("taskkill /f /im  EXCEL.EXE")
        

class SaveToFile():
    def __init__(self, master):
        
        self.master = master
        
        if os.path.isfile(self.master) is True:
            pass
        else:
            wb = Workbook()
            ws = wb.active
            
            ws.cell(row=2, column=2, value="Varenummer:")
            ws.cell(row=2, column=3, value="Bearbeidningstid:")
            
            wb.save(self.master)
            
        self.tablewrite()
            
    def tablewrite(self):
        
        """ Writing the results to a excel file """
        
        wb = openpyxl.load_workbook(self.master)
        ws = wb.active
        
        row_count = 3
        row = 3
        
        p_number = TimeCalc(readfile).pnumber()
        time = TimeCalc(readfile).results()
        count_var = 0
        
        while ws.cell(row=row_count, column=2).value != None:
            if ws.cell(row=row_count, column=2).value == p_number:
                ws.cell(row=row, column=2, value=p_number)
                ws.cell(row=row, column=3, value=time)                
                count_var = 1
                            
            row_count += 1
            row += 1
        
        if count_var == 0:
            ws.cell(row=row, column=2, value=p_number)
            ws.cell(row=row, column=3, value=time)
            
        # centering the written results in the excel file 
        cell_al1 = ws.cell(row=row, column=2)
        cell_al1.alignment = Alignment(horizontal='center')
        
        cell_al2 = ws.cell(row=row, column=3)
        cell_al2.alignment = Alignment(horizontal='center')
        
        wb.save(self.master)
            
            
readfile = os.path.expanduser("~\\Documents\\report.xlsx")
savefile = "O:\\Maskin 20\\Tidskalkyle\\Tidskalkyle.xlsx"
     
SaveToFile(savefile)




#input("Press Enter to continue...")


        